import datetime
import glob
import os
import shutil
import sys
import time
import zipfile

import numpy as np
import openpyxl
import pandas as pd
from pandas._libs.index import timedelta

import xlrd
import pytz

from pathlib import Path
from datetime import datetime
from datetime import timezone



sys.path.append(os.path.join(os.path.dirname(__file__), "...", "..."))
p = Path(os.getcwd())
dirctorypath = p.parent.parent
testdatafolder = "testData"
performancetestdatafolder = "testData\\Performance_Circuits"
logfilefolder = "src\\Reports\\logs\\"
logfilefolder_mac = "src/Reports/logs/"
testDatafile = "testData.xlsx"

PerformanceDatafolderPath = os.path.join(dirctorypath, performancetestdatafolder)
testDatafolderPath = os.path.join(dirctorypath, testdatafolder)
testDatafilePath = os.path.join(testDatafolderPath, testDatafile)
logfilepath = os.path.join(dirctorypath, logfilefolder)
downloadsfolder = os.path.join(p.parent.parent, "downloads/")
downloadsfolderPath = os.path.join(dirctorypath, downloadsfolder)
driver = None


# Method: readData
# Method Desc: Read Excel file
# input : file, sheetName, rownum, columnno
# return: cell value


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value


# Method: getRowCount
# Method Desc: get total row count of Excel file
# input : filepath & fileName
# return: Total Count
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


# Method: deleteFiles
# Method Desc: Delete files from folder
# input : filePath,fileType
# return: none
def deleteFiles(filePath, fileType):
    try:
        for file in os.scandir(filePath):
            if file.name.endswith(fileType):
                os.unlink(file.path)
    except:
        None


# Method: getCSVrowCount
# Method Desc: get total row count of CSV file
# input : filepath & fileName
# return: Total Count
def getCSVrowCount(filepath, fileName):
    file = os.path.join(filepath, fileName)
    rCount = pd.read_csv(file)
    return len(rCount)


# Method: convertExcelToText
# Method Desc: Coverts data from Excel to Text file
# input : file, index, TextFileName
# return: none
def convertExcelToTextIndex(file, index, TextFileName):
    workbook = xlrd.open_workbook(file, logfile=open(os.devnull, 'w'))
    sheet = workbook.sheet_by_index(index)
    Total_cols = sheet.ncols
    Total_rows = sheet.nrows
    for row_index in range(0, Total_rows):
        data = []
        for row_index in range(0, Total_rows):
            row = []
            for col_index in range(0, Total_cols):
                var_cell_value = str(sheet.cell(row_index, col_index).value)
                if var_cell_value.find('.0') >= 0:
                    var_cell_value = var_cell_value.split('.')[0]
                row.append(var_cell_value)
            data.append(row)
            with open(TextFileName, 'w') as f:
                f.write("\n".join(','.join(map(str, row)) for row in data))


# Method: convertTexttoCSV
# Method Desc: Coverts data from text to CSV file
# input : textFile,csvFile
# return: none
def convertTexttoCSV(textFile, csvFile):
    if os.path.isfile(csvFile):
        os.remove(csvFile)
    os.rename(textFile, csvFile)


# Method to get most recent file from download folder
# Returns most recent file.
def getMostRecent_downloaded_File(ext_pattern=None):
    if ext_pattern is not None:
        pattern = ext_pattern
    else:
        pattern = "*.csv"
    time.sleep(1)
    try:
        # wait_tillfiledownloads()
        list_of_files = glob.glob(downloadsfolderPath + pattern)
        latest_file = max(list_of_files, key=os.path.getctime)
        print(latest_file)
        var_mostRecentFile = os.path.join(downloadsfolderPath, latest_file)
        return var_mostRecentFile
    except:
        wait_tillfiledownloads()
        list_of_files = glob.glob(downloadsfolderPath + pattern)

        latest_file = max(list_of_files, key=os.path.getctime)
        print(latest_file)
        var_mostRecentFile = os.path.join(downloadsfolderPath, latest_file)
        return var_mostRecentFile


def wait_tillfiledownloads():
    var_fileextn = "crdownload"
    time.sleep(2)
    while "crdownload" in var_fileextn:
        for var_recenfile in os.listdir(downloadsfolderPath):
            print(var_recenfile)
            if "crdownload" in var_recenfile:
                var_fileextn = "crdownload"
                time.sleep(3)
            else:
                var_fileextn = "None"
                time.sleep(3)
    time.sleep(3)


def getCurrentTime():
    var_now = datetime.now()
    return (str(var_now.strftime("%Y%m%d_%H%M%S")))


# Method: deleteFiles
# Method Desc: Delete files from folder
# input : filePath,fileType
# return: none
def deleteFolder(foldername):
    if os.path.exists(foldername) and os.path.isdir(foldername):
        shutil.rmtree(foldername)


# Create directory
def create_folder(dir):
    os.mkdir(dir)


def comparedf(df1, df2):
    """Identify differences between two pandas DataFrames"""
    assert (df1.columns == df2.columns).all(), \
        "DataFrame column names are different"
    if any(df1.dtypes != df2.dtypes):
        "Data Types are different, trying to convert"
        df2 = df2.astype(df1.dtypes)
    if df1.equals(df2):
        return True, "both DataFrame are equal"
    else:
        # need to account for np.nan != np.nan returning True
        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
        ne_stacked = diff_mask.stack()
        changed = ne_stacked[ne_stacked]
        changed.index.names = ['id', 'col']
        difference_locations = np.where(diff_mask)
        changed_from = df1.values[difference_locations]
        changed_to = df2.values[difference_locations]
        return pd.DataFrame({'from': changed_from, 'to': changed_to},
                            index=changed.index)


def timezoneconversion_utc(df_datecolumn):
    columns = []
    for date_str in df_datecolumn:
        columns.append(convert_datetime(date_str))
    return columns


def convert_datetime(x):
    """This function is convert the time to PST & then UTC time zone

        Args:
            x - column values
        Returns:
            return date in UTC time zone otherwise return Not a Time
    """
    try:
        pst = pytz.timezone('US/Pacific')
        dt = datetime.strptime(x, '%Y-%m-%d %H:%M:%S')
        dt_pst = pst.localize(dt)
        dt_utc = datetime.fromtimestamp(dt_pst.timestamp(), timezone.utc)
        return 'T'.join(str(datetime.strptime(str(dt_utc).split('+')[0], '%Y-%m-%d %H:%M:%S')).split(' '))
    except:
        return None


# Method: compare_two_dataframe_coulumns
# Method Desc: Match two data frames
# input : dataframe1 & dataframe2
def compare_two_columns_dataframe(df1, df2):
    arr_mismatchdata = []
    try:
        for (i, j) in zip(df1, df2):
            if i != j:
                arr_mismatchdata.append([i, j])
    except:
        assert False
    return arr_mismatchdata


# Unzip a file to folder
def unzip_file(filename, extract_to_directory):
    with zipfile.ZipFile(filename, "r") as zip_ref:
        zip_ref.extractall(extract_to_directory)
        count = len(zip_ref.infolist())
        print(count)
        files = zip_ref.namelist()
        print(files)
    return count, files

# def metaData_modal_Reports_with_flag_and_calender(self):
#     try:
#         date_value_today = getcurrentTimevalue()
#         date_value = getTimeAddedvalue(day_count)
#         self.Click(locators.md_estshutoffstarttime_date_txt)
#
#         self.setText(getTimeAddedvalue(day_count), locators.md_estshutoffstarttime_date_txt)
#         print("Clicked Estimated Shut Off Start Time")
#         self.Click(locators.md_estshutoffendtime_date_txt)
#         self.setText(getTimeAddedvalue(day_count, 1), locators.md_estshutoffendtime_date_txt)
#         print("Clicked Estimated Shut Off end Time")
#         self.Click(locators.md_allclear_date_txt)
#         self.setText(getTimeAddedvalue(day_count, 2), locators.md_allclear_date_txt)
#         print("Clicked All clear")
#         self.Click(locators.md_etor_date_txt)
#         self.setText(getTimeAddedvalue(day_count, 3), locators.md_etor_date_txt)
#         print("Clicked ETOR")
#         self.setText("Automation", locators.md_input_Comments)
#         print("Entered Comments:  ")
#         self.select_flag(flag_value)
#         print("Selected Flag: %s " % str(flag_value))
#         time.sleep(2)
#         self.Click(locators.md_btn_Save)
#         print("Clicked on Save meta data button")
#     except(ValueError, Exception):
#         return False
#     return True

# Method: getCurrentTime
# Method Desc: get current system time/ PST
# return: Time
def getTimeAddedvalue(day_count, hr_count=0):
    d = datetime.datetime.today() + timedelta(days=int(day_count), hours=int(hr_count))
    return (str(d.strftime("%m%d%Y%H%M")))




